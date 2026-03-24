from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.db.models import Q
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.conf import settings as django_settings
from datetime import date, datetime, timedelta
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import (Student, Attendance, Teacher, ClassSection, Subject,
                     Holiday, LeaveRequest, ActivityLog, Notification, Timetable)

PASSWORD_MIN_MSG = "Password must be at least 6 characters."

def is_teacher(request):  return request.session.get('role') == 'teacher'
def is_student(request):  return request.session.get('role') == 'student'
def is_admin(request):    return request.session.get('role') == 'teacher' and request.session.get('is_admin', False)

def get_greeting():
    h = datetime.now().hour
    if h < 12: return "Good Morning"
    elif h < 17: return "Good Afternoon"
    elif h < 21: return "Good Evening"
    return "Good Night"

def teacher_required(fn):
    @wraps(fn)
    def wrapper(request, *args, **kwargs):
        if not is_teacher(request):
            messages.error(request, "Please login as teacher.")
            return redirect('login')
        return fn(request, *args, **kwargs)
    return wrapper

def admin_required(fn):
    @wraps(fn)
    def wrapper(request, *args, **kwargs):
        if not is_admin(request):
            messages.error(request, "Admin access required.")
            return redirect('home')
        return fn(request, *args, **kwargs)
    return wrapper

def student_attendance_stats(student):
    total   = Attendance.objects.filter(student=student).count()
    present = Attendance.objects.filter(student=student, status='Present').count()
    absent  = total - present
    pct     = round((present / total) * 100, 1) if total > 0 else 0.0
    return present, absent, total, pct

def attendance_prediction(student):
    """Return how many classes student can miss before dropping below 75%."""
    present, absent, total, pct = student_attendance_stats(student)
    threshold = getattr(django_settings, 'ATTENDANCE_THRESHOLD', 75)
    if total == 0: return None, pct
    # Solve: present/(total+x) >= 0.75  =>  x <= present/0.75 - total
    can_miss = int(present / (threshold / 100)) - total
    return max(0, can_miss), pct

def log_activity(request, action, message):
    role = request.session.get('role', 'unknown')
    if role == 'teacher':
        actor = request.session.get('teacher_name', 'Teacher')
    else:
        actor = request.session.get('student_name', 'Student')
    ip = request.META.get('REMOTE_ADDR', '')
    ActivityLog.objects.create(actor=actor, role=role, action=action, message=message, ip_address=ip)

def push_notification(title, message, ntype='info', student=None, teacher=None):
    Notification.objects.create(
        title=title, message=message, type=ntype,
        recipient_student=student, recipient_teacher=teacher
    )

def get_unread_count(request):
    if is_teacher(request):
        tid = request.session.get('teacher_db_id')
        if tid:
            return Notification.objects.filter(recipient_teacher_id=tid, is_read=False).count()
    elif is_student(request):
        sid = request.session.get('student_id')
        if sid:
            return Notification.objects.filter(recipient_student_id=sid, is_read=False).count()
    return 0


def parse_calendar_month(month_param, default_date):
    if not month_param:
        return default_date.year, default_date.month
    try:
        year_str, month_str = month_param.split('-', 1)
        year = int(year_str)
        month = int(month_str)
        if 1 <= month <= 12:
            return year, month
    except (ValueError, TypeError):
        pass
    return default_date.year, default_date.month


def build_student_calendar_data(student, month_param, today=None):
    today = today or date.today()
    cal_year, cal_month = parse_calendar_month(month_param, today)

    first_day = date(cal_year, cal_month, 1)
    if cal_month == 12:
        last_day = date(cal_year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(cal_year, cal_month + 1, 1) - timedelta(days=1)

    prev_year = cal_year - 1 if cal_month == 1 else cal_year
    prev_month = 12 if cal_month == 1 else cal_month - 1
    next_year = cal_year + 1 if cal_month == 12 else cal_year
    next_month = 1 if cal_month == 12 else cal_month + 1

    month_att = {}
    month_records = Attendance.objects.filter(
        student=student,
        date__year=cal_year,
        date__month=cal_month,
    ).only('date', 'status')
    for a in month_records:
        existing = month_att.get(a.date)
        if existing == 'Absent':
            continue
        if a.status == 'Absent':
            month_att[a.date] = 'Absent'
        elif existing is None:
            month_att[a.date] = 'Present'
    month_holidays = {
        h.date: h.name
        for h in Holiday.objects.filter(date__year=cal_year, date__month=cal_month)
    }

    calendar_blank = list(range(first_day.weekday()))
    calendar_days = []
    d = first_day
    while d <= last_day:
        status = month_att.get(d)
        holiday = month_holidays.get(d)
        calendar_days.append({
            'date': d,
            'status': status,
            'holiday': holiday,
            'is_today': d == today,
        })
        d += timedelta(days=1)

    return {
        'cal_month': first_day.strftime('%B %Y'),
        'cal_month_value': f"{cal_year:04d}-{cal_month:02d}",
        'prev_cal_month': f"{prev_year:04d}-{prev_month:02d}",
        'next_cal_month': f"{next_year:04d}-{next_month:02d}",
        'calendar_blank': calendar_blank,
        'calendar_days': calendar_days,
    }


def login_page(request):
    if is_teacher(request): return redirect('home')
    if is_student(request):  return redirect('student_dashboard')
    return render(request, "attendance/login.html")

def teacher_login(request):
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        # Check DB teachers first
        try:
            teacher = Teacher.objects.get(username=username)
            if teacher.check_password(password):
                request.session.flush()
                request.session['role']          = 'teacher'
                request.session['teacher_name']  = teacher.name
                request.session['teacher_db_id'] = teacher.id
                request.session['is_admin']       = teacher.is_admin
                log_activity(request, 'login', f"Teacher '{teacher.name}' logged in")
                return redirect('home')
            else:
                messages.error(request, "Invalid username or password.")
                return render(request, "attendance/teacher_login.html")
        except Teacher.DoesNotExist:
            pass
        # Fallback to .env credentials
        valid_user = getattr(django_settings, 'TEACHER_USERNAME', 'admin')
        valid_pass = getattr(django_settings, 'TEACHER_PASSWORD', 'Admin@1234')
        if username == valid_user and password == valid_pass:
            request.session.flush()
            request.session['role']         = 'teacher'
            request.session['teacher_name'] = 'Admin'
            request.session['is_admin']      = True
            log_activity(request, 'login', "Admin logged in via .env credentials")
            return redirect('home')
        messages.error(request, "Invalid username or password.")
    return render(request, "attendance/teacher_login.html")

def student_login(request):
    if request.method == "POST":
        email    = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        try:
            student = Student.objects.get(email=email)
            if student.check_password(password):
                request.session['student_id']   = student.id
                request.session['role']          = 'student'
                request.session['student_name']  = student.name
                log_activity(request, 'login', f"Student '{student.name}' logged in")
                return redirect('student_dashboard')
            messages.error(request, "Invalid email or password.")
        except Student.DoesNotExist:
            messages.error(request, "Invalid email or password.")
    return render(request, 'attendance/student_login.html')

def forgot_password(request):
    if request.method == "POST":
        email = request.POST.get("email", "").strip()
        np    = request.POST.get("new_password", "")
        cp    = request.POST.get("confirm_password", "")
        if np != cp:
            messages.error(request, "Passwords do not match.")
        elif len(np) < 6:
            messages.error(request, PASSWORD_MIN_MSG)
        else:
            try:
                student = Student.objects.get(email=email)
                student.set_password(np); student.save()
                messages.success(request, "Password reset successfully!")
                return redirect("student_login")
            except Student.DoesNotExist:
                messages.error(request, "No account found with that email.")
    return render(request, "attendance/forgot_password.html")

def teacher_forgot_password(request):
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        try:
            teacher = Teacher.objects.get(username=username)
            messages.success(request, f"Contact admin to reset password for '{username}'.")
        except Teacher.DoesNotExist:
            valid_user = getattr(django_settings, 'TEACHER_USERNAME', 'admin')
            if username == valid_user:
                messages.success(request, "Update TEACHER_PASSWORD in your .env file.")
            else:
                messages.error(request, "Username not found.")
    return render(request, "attendance/teacher_forgot_password.html")

def logout_view(request):
    log_activity(request, 'logout', f"{request.session.get('teacher_name') or request.session.get('student_name','?')} logged out")
    request.session.flush()
    return redirect("login")


def legacy_holidays_redirect(request, *args, **kwargs):
    messages.info(request, "Holiday calendar has been removed.")
    if is_teacher(request):
        return redirect('home')
    if is_student(request):
        return redirect('student_dashboard')
    return redirect('login')


@teacher_required
def home(request):
    selected_date = request.GET.get('date')
    filter_date   = datetime.strptime(selected_date, '%Y-%m-%d').date() if selected_date else date.today()
    is_holiday    = Holiday.objects.filter(date=filter_date).first()

    total_students   = Student.objects.count()
    present_today    = Attendance.objects.filter(date=filter_date, status="Present", subject=None).count()
    absent_today     = Attendance.objects.filter(date=filter_date, status="Absent",  subject=None).count()
    pending_leaves   = LeaveRequest.objects.filter(status='Pending').count()
    unread           = get_unread_count(request)

    names, percentages, statuses = [], [], []
    for student in Student.objects.all():
        att    = Attendance.objects.filter(student=student, date=filter_date, subject=None).first()
        status = att.status if att else "None"
        names.append(student.name); statuses.append(status)
        percentages.append(100 if status == "Present" else 0)

    return render(request, "attendance/dashboard.html", {
        "total_students": total_students, "present_today": present_today,
        "absent_today": absent_today, "today_attendance": present_today + absent_today,
        "pending_leaves": pending_leaves, "unread": unread,
        "names": names, "percentages": percentages, "statuses": statuses,
        "selected_date": str(filter_date), "is_holiday": is_holiday,
        "greeting": get_greeting(), "teacher_name": request.session.get('teacher_name', 'Teacher'),
    })


@teacher_required
def bulk_import_students(request):
    if request.method == "POST":
        uploaded = request.FILES.get("file")
        if not uploaded:
            messages.error(request, "Please upload a file.")
            return redirect("bulk_import_students")

        fname = uploaded.name.lower()
        added = 0; skipped = 0; errors = []

        try:
            if fname.endswith(".csv"):
                import csv, io
                decoded = uploaded.read().decode("utf-8", errors="ignore")
                reader  = csv.DictReader(io.StringIO(decoded))
                rows    = list(reader)
            elif fname.endswith((".xlsx", ".xls")):
                import openpyxl
                wb   = openpyxl.load_workbook(uploaded)
                ws   = wb.active
                hdrs = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({hdrs[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
            else:
                messages.error(request, "Only .csv or .xlsx files are supported.")
                return redirect("bulk_import_students")

            for i, row in enumerate(rows, start=2):
                # Accept flexible column names
                name  = row.get("name") or row.get("student name") or row.get("full name") or ""
                roll  = row.get("roll") or row.get("roll number") or row.get("roll_number") or row.get("rollno") or ""
                email = row.get("email") or row.get("email address") or ""
                pwd   = row.get("password") or row.get("pass") or "12345"
                cs    = row.get("class") or row.get("class_section") or ""

                name = name.strip(); roll = roll.strip(); email = email.strip()

                if not name or not roll:
                    errors.append(f"Row {i}: Name and Roll are required.")
                    skipped += 1
                    continue

                # Auto-generate email if missing
                if not email:
                    email = roll.lower().replace(" ", "") + "@student.edu"

                # Skip duplicates
                if Student.objects.filter(roll_number=roll).exists():
                    skipped += 1
                    continue
                if Student.objects.filter(email=email).exists():
                    email = roll.lower() + str(i) + "@student.edu"

                # Find class section if provided
                cs_obj = None
                if cs:
                    cs_obj = ClassSection.objects.filter(name__iexact=cs).first()

                student = Student(roll_number=roll, name=name, email=email, class_section=cs_obj)
                student.set_password(str(pwd) if pwd else "12345")
                student.save()
                _send_welcome_email(student)
                added += 1

        except Exception as e:
            messages.error(request, f"Error reading file: {e}")
            return redirect("bulk_import_students")

        log_activity(request, "add", f"Bulk imported {added} students from file")
        if added:
            messages.success(request, f"✅ {added} student(s) imported successfully!{' ' + str(skipped) + ' skipped.' if skipped else ''}")
        else:
            messages.warning(request, f"No students imported. {skipped} skipped (duplicates or missing data).")
        if errors:
            for e in errors[:5]:
                messages.error(request, e)
        return redirect("students")

    # GET — show upload form
    classes = ClassSection.objects.all()
    return render(request, "attendance/bulk_import.html", {"classes": classes})


def student_register(request):
    """Students can register themselves — teacher approves later."""
    if request.method == "POST":
        name  = request.POST.get("name","").strip()
        roll  = request.POST.get("roll_number","").strip()
        email = request.POST.get("email","").strip()
        pwd   = request.POST.get("password","")
        cpwd  = request.POST.get("confirm_password","")
        cs_id = request.POST.get("class_section","")

        errors = []
        if not name:  errors.append("Full name is required.")
        if not roll:  errors.append("Roll number is required.")
        if not email: errors.append("Email is required.")
        if len(pwd) < 6: errors.append(PASSWORD_MIN_MSG)
        if pwd != cpwd:  errors.append("Passwords do not match.")
        if Student.objects.filter(roll_number=roll).exists(): errors.append(f"Roll number '{roll}' already registered.")
        if Student.objects.filter(email=email).exists():      errors.append(f"Email '{email}' already registered.")

        if errors:
            for e in errors: messages.error(request, e)
            return render(request, "attendance/student_register.html", {
                "classes": ClassSection.objects.all(), "form_data": request.POST
            })

        student = Student(roll_number=roll, name=name, email=email)
        student.set_password(pwd)
        if cs_id: student.class_section_id = cs_id
        if request.FILES.get("photo"): student.photo = request.FILES["photo"]
        student.save()
        log_activity(request, "add", f"Student '{name}' self-registered")
        messages.success(request, "Registration successful! You can now login.")
        return redirect("student_login")

    return render(request, "attendance/student_register.html", {
        "classes": ClassSection.objects.all()
    })

@teacher_required
def add_student(request):
    classes = ClassSection.objects.all()
    if request.method == "POST":
        roll  = request.POST.get('roll_number','').strip()
        name  = request.POST.get('name','').strip()
        email = request.POST.get('email','').strip()
        pwd   = request.POST.get('password','')
        cs_id = request.POST.get('class_section')
        phone = request.POST.get('phone','').strip()
        pphone= request.POST.get('parent_phone','').strip()
        gender= request.POST.get('gender','')

        errors = []
        if not roll:  errors.append("Roll number is required.")
        if not name:  errors.append("Name is required.")
        if not email: errors.append("Email is required.")
        if len(pwd) < 6: errors.append(PASSWORD_MIN_MSG)
        if Student.objects.filter(roll_number=roll).exists(): errors.append(f"Roll '{roll}' already exists.")
        if Student.objects.filter(email=email).exists():      errors.append(f"Email '{email}' already registered.")

        if errors:
            for e in errors: messages.error(request, e)
            return render(request, 'attendance/add_student.html', {'classes': classes, 'form_data': request.POST})

        student = Student(roll_number=roll, name=name, email=email, phone=phone, parent_phone=pphone, gender=gender)
        student.set_password(pwd)
        if cs_id: student.class_section_id = cs_id
        if request.FILES.get('photo'): student.photo = request.FILES['photo']
        student.save()
        _send_welcome_email(student)
        log_activity(request, 'add', f"Added student '{name}' ({roll})")
        push_notification(f"Welcome, {name}!", "Your account has been created. Login to view your attendance.", 'success', student=student)
        messages.success(request, f"Student '{name}' added successfully!")
        return redirect('students')
    return render(request, 'attendance/add_student.html', {'classes': classes})

def _send_welcome_email(student):
    if not student.email: return
    try:
        send_mail("Welcome to EduTrack 🎓",
            f"Dear {student.name},\n\nYour account has been created.\n\nLogin: {student.email}\nRoll No: {student.roll_number}\n\nRegards,\nEduTrack",
            django_settings.DEFAULT_FROM_EMAIL, [student.email], fail_silently=True)
    except Exception:
        pass

@teacher_required
def students(request):
    query      = request.GET.get("q","")
    class_id   = request.GET.get("class_id","")
    page_num   = request.GET.get("page",1)
    qs         = Student.objects.select_related('class_section').all()
    if query:    qs = qs.filter(Q(name__icontains=query)|Q(roll_number__icontains=query))
    if class_id: qs = qs.filter(class_section_id=class_id)

    suggestion_rows = Student.objects.order_by('name').values_list('name', 'roll_number')[:500]
    student_suggestions = []
    seen = set()
    for name, roll in suggestion_rows:
        key = (name or '').strip()
        if key and key.lower() not in seen:
            seen.add(key.lower())
            initial = key[0].upper() if key else '?'
            student_suggestions.append({
                'name': key,
                'roll': (roll or '').strip(),
                'initial': initial,
                'display': f"{key} ({roll})" if roll else key
            })

    paginator  = Paginator(qs, 10)
    page_obj   = paginator.get_page(page_num)
    return render(request, "attendance/students.html", {
        "students": page_obj, "page_obj": page_obj,
        "query": query, "class_id": class_id,
        "classes": ClassSection.objects.all(),
        "student_suggestions": student_suggestions,
    })

def edit_student(request, id):
    if is_teacher(request):
        student     = get_object_or_404(Student, id=id); redirect_to = "students"
    elif is_student(request):
        sid = request.session.get('student_id')
        if str(sid) != str(id):
            messages.error(request, "You can only edit your own profile."); return redirect('student_dashboard')
        student = get_object_or_404(Student, id=sid); redirect_to = "student_dashboard"
    else: return redirect('login')

    if request.method == "POST":
        if request.GET.get("crop") == "save" and request.headers.get("X-Requested-With") == "XMLHttpRequest":
            cropped_image = request.FILES.get("cropped_image")
            if not cropped_image:
                return JsonResponse({"ok": False, "error": "No cropped image provided."}, status=400)
            student.photo = cropped_image
            student.save(update_fields=["photo"])
            log_activity(request, 'edit', f"Cropped profile photo for '{student.name}'")
            return JsonResponse({"ok": True})

        name  = request.POST.get("name","").strip()
        roll  = request.POST.get("roll","").strip()
        email = request.POST.get("email","").strip()
        if Student.objects.filter(roll_number=roll).exclude(id=id).exists():
            messages.error(request, f"Roll '{roll}' already taken."); return render(request,"attendance/edit_student.html",{"student":student,"classes":ClassSection.objects.all()})
        if Student.objects.filter(email=email).exclude(id=id).exists():
            messages.error(request, f"Email '{email}' already registered."); return render(request,"attendance/edit_student.html",{"student":student,"classes":ClassSection.objects.all()})
        student.name = name; student.roll_number = roll; student.email = email
        student.phone = request.POST.get("phone","").strip()
        student.parent_phone = request.POST.get("parent_phone","").strip()
        student.gender = request.POST.get("gender","")
        cs_id = request.POST.get("class_section")
        student.class_section_id = cs_id if cs_id else None
        if request.FILES.get("photo"): student.photo = request.FILES["photo"]
        if request.POST.get("remove_photo"): student.photo.delete(save=False); student.photo = None
        np = request.POST.get("new_password","")
        if np:
            if len(np) < 6:
                messages.error(request,"Password must be 6+ characters."); return render(request,"attendance/edit_student.html",{"student":student,"classes":ClassSection.objects.all()})
            student.set_password(np)
        student.save()
        log_activity(request, 'edit', f"Edited student '{student.name}'")
        messages.success(request, "Profile updated!"); return redirect(redirect_to)
    return render(request,"attendance/edit_student.html",{"student":student,"classes":ClassSection.objects.all()})

@teacher_required
def delete_student(request, id):
    s = get_object_or_404(Student, id=id); name = s.name; s.delete()
    log_activity(request, 'delete', f"Deleted student '{name}'")
    messages.success(request, f"Student '{name}' deleted."); return redirect("students")


@admin_required
def teachers(request):
    teacher_list = Teacher.objects.all()
    return render(request, "attendance/teachers.html", {"teachers": teacher_list})

@admin_required
def add_teacher(request):
    classes = ClassSection.objects.all()
    if request.method == "POST":
        name  = request.POST.get('name','').strip()
        uname = request.POST.get('username','').strip()
        email = request.POST.get('email','').strip()
        pwd   = request.POST.get('password','')
        is_ad = request.POST.get('is_admin') == 'on'
        errors = []
        if not name:  errors.append("Name required.")
        if not uname: errors.append("Username required.")
        if len(pwd) < 6: errors.append("Password must be 6+ characters.")
        if Teacher.objects.filter(username=uname).exists(): errors.append(f"Username '{uname}' taken.")
        if email and Teacher.objects.filter(email=email).exists(): errors.append(f"Email '{email}' taken.")
        if errors:
            for e in errors: messages.error(request,e)
            return render(request,'attendance/add_teacher.html',{'classes':classes,'form_data':request.POST})
        t = Teacher(name=name, username=uname, email=email or None, is_admin=is_ad)
        t.set_password(pwd)
        if request.FILES.get('photo'): t.photo = request.FILES['photo']
        t.save()
        log_activity(request,'add',f"Added teacher '{name}'")
        messages.success(request,f"Teacher '{name}' added!"); return redirect('teachers')
    return render(request,'attendance/add_teacher.html',{'classes':classes})

@admin_required
def edit_teacher(request, id):
    teacher = get_object_or_404(Teacher, id=id)
    if request.method == "POST":
        teacher.name     = request.POST.get('name','').strip()
        teacher.email    = request.POST.get('email','').strip() or None
        teacher.is_admin = request.POST.get('is_admin') == 'on'
        np = request.POST.get('new_password','')
        if np:
            if len(np) < 6: messages.error(request,"Password 6+ chars"); return render(request,'attendance/add_teacher.html',{'teacher':teacher,'edit':True})
            teacher.set_password(np)
        if request.FILES.get('photo'): teacher.photo = request.FILES['photo']
        teacher.save(); messages.success(request,"Teacher updated!"); return redirect('teachers')
    return render(request,'attendance/add_teacher.html',{'teacher':teacher,'edit':True})

@admin_required
def delete_teacher(request, id):
    t = get_object_or_404(Teacher,id=id); name=t.name; t.delete()
    messages.success(request,f"Teacher '{name}' deleted."); return redirect('teachers')


@teacher_required
def classes(request):
    return render(request,"attendance/classes.html",{"classes":ClassSection.objects.prefetch_related('students').all(),"teachers":Teacher.objects.all()})

@teacher_required
def add_class(request):
    if request.method == "POST":
        name    = request.POST.get('name','').strip()
        section = request.POST.get('section','').strip()
        t_id    = request.POST.get('teacher')
        if not name: messages.error(request,"Class name required."); return redirect('classes')
        ClassSection.objects.create(name=name, section=section, teacher_id=t_id or None)
        messages.success(request,f"Class '{name}' created!"); return redirect('classes')
    return redirect('classes')

@teacher_required
def delete_class(request, id):
    c = get_object_or_404(ClassSection,id=id); c.delete()
    messages.success(request,"Class deleted."); return redirect('classes')

@teacher_required
def subjects(request):
    subjects_qs = (
        Subject.objects
        .select_related('class_section', 'teacher')
        .order_by('class_section__name', 'class_section__section', 'name')
    )
    classes_qs = ClassSection.objects.order_by('name', 'section')
    teachers_qs = Teacher.objects.order_by('name')

    return render(request, "attendance/subjects.html", {
        "subjects": subjects_qs,
        "classes": classes_qs,
        "teachers": teachers_qs,
    })

@teacher_required
def add_subject(request):
    if request.method == "POST":
        name  = request.POST.get('name','').strip()
        cs_id = request.POST.get('class_section')
        t_id  = request.POST.get('teacher')
        if not name: messages.error(request,"Subject name required."); return redirect('subjects')
        Subject.objects.create(name=name, class_section_id=cs_id or None, teacher_id=t_id or None)
        messages.success(request,f"Subject '{name}' added!"); return redirect('subjects')
    return redirect('subjects')

@teacher_required
def delete_subject(request, id):
    s = get_object_or_404(Subject,id=id); s.delete()
    messages.success(request,"Subject deleted."); return redirect('subjects')


@teacher_required
def mark_page(request):
    today      = date.today()
    is_holiday = Holiday.objects.filter(date=today).first()
    lecture_id = request.GET.get('lecture_id', '')
    today_day = today.strftime('%A')
    lectures_today = Timetable.objects.filter(day=today_day).select_related('class_section', 'subject', 'teacher').order_by('start_time', 'class_section__name', 'class_section__section')

    selected_lecture = None
    if lecture_id:
        selected_lecture = lectures_today.filter(id=lecture_id).first()
    if not selected_lecture and lectures_today.exists():
        selected_lecture = lectures_today.first()
        lecture_id = str(selected_lecture.id)

    selected_subject = selected_lecture.subject if selected_lecture else None

    if selected_lecture:
        students = Student.objects.filter(class_section=selected_lecture.class_section).order_by('roll_number')
    else:
        students = Student.objects.none()

    for student in students:
        att = Attendance.objects.filter(
            student=student,
            date=today,
            subject=selected_subject,
            timetable_entry=selected_lecture,
        ).first()
        student.today_status = att.status if att else None
        student.today_note   = att.note   if att else ''

    present_today = Attendance.objects.filter(date=today, status='Present', subject=selected_subject, timetable_entry=selected_lecture).count()
    absent_today  = Attendance.objects.filter(date=today, status='Absent',  subject=selected_subject, timetable_entry=selected_lecture).count()
    pending_today = students.count() - present_today - absent_today

    return render(request, 'attendance/mark_attendance.html', {
        'students': students, 'today': today,
        'present_today': present_today, 'absent_today': absent_today, 'pending_today': pending_today,
        'is_holiday': is_holiday,
        'selected_subject': selected_subject,
        'lectures_today': lectures_today,
        'selected_lecture': selected_lecture,
        'lecture_id': lecture_id,
        'today_day': today_day,
    })

@teacher_required
def mark_attendance(request, student_id, status):
    student  = get_object_or_404(Student, id=student_id)
    today    = date.today()
    lecture_id = request.POST.get('lecture_id') or request.GET.get('lecture_id')
    note       = request.POST.get('note','')
    lecture    = None
    subject    = None
    if lecture_id:
        lecture = get_object_or_404(Timetable, id=lecture_id)
        subject = lecture.subject

        if student.class_section_id != lecture.class_section_id:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return HttpResponse(status=400)
            messages.error(request, f"{student.name} does not belong to {lecture.class_section}.")
            return redirect(f"/mark/?lecture_id={lecture.id}")

    att, created = Attendance.objects.get_or_create(student=student, date=today, subject=subject, timetable_entry=lecture, defaults={'status':status,'note':note})
    if not created: att.status=status; att.note=note; att.save()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest': return HttpResponse(status=200)
    return redirect(f"/mark/?lecture_id={lecture_id}" if lecture_id else '/mark/')

@teacher_required
def edit_attendance(request, attendance_id):
    att = get_object_or_404(Attendance, id=attendance_id)
    if request.method == "POST":
        ns = request.POST.get("status")
        if ns in ['Present','Absent']: att.status=ns; att.note=request.POST.get("note",""); att.save()
        messages.success(request, f"Attendance updated for {att.student.name} on {att.date}.")
    return redirect(request.POST.get('next','report'))

@teacher_required
def bulk_mark_present(request):
    today = date.today()
    lecture_id = request.POST.get('lecture_id')
    lecture = None
    subject = None
    students_qs = Student.objects.all()

    if lecture_id:
        lecture = get_object_or_404(Timetable, id=lecture_id)
        subject = lecture.subject
        students_qs = students_qs.filter(class_section=lecture.class_section)

    count = 0
    for student in students_qs:
        obj, created = Attendance.objects.get_or_create(student=student, date=today, subject=subject, timetable_entry=lecture, defaults={'status':'Present'})
        if not created and obj.status != 'Present': obj.status='Present'; obj.save()
        count += 1
    log_activity(request,'attendance',f"Bulk marked all {count} students Present")
    messages.success(request, f"All {count} students marked Present!")
    return redirect(f"/mark/?lecture_id={lecture_id}" if lecture_id else 'mark_page')


@teacher_required
def report(request):
    query      = request.GET.get('q','')
    date_from  = request.GET.get('date_from','')
    date_to    = request.GET.get('date_to','')
    class_id   = request.GET.get('class_id','')
    subject_id = request.GET.get('subject_id','')
    threshold  = getattr(django_settings,'ATTENDANCE_THRESHOLD',75)

    qs = Student.objects.all()
    if query:    qs = qs.filter(Q(name__icontains=query)|Q(roll_number__icontains=query))
    if class_id: qs = qs.filter(class_section_id=class_id)

    selected_subject = None
    if subject_id:
        try: selected_subject = Subject.objects.get(id=subject_id)
        except Subject.DoesNotExist:
            pass

    report_data = []
    for student in qs:
        att = Attendance.objects.filter(student=student)
        if selected_subject: att = att.filter(subject=selected_subject)
        if date_from: att = att.filter(date__gte=date_from)
        if date_to:   att = att.filter(date__lte=date_to)
        present = att.filter(status='Present').count()
        absent  = att.filter(status='Absent').count()
        total   = present + absent
        pct     = round((present/total*100),1) if total>0 else 0
        report_data.append({'student':student,'present':present,'absent':absent,'total':total,'percentage':pct,'low':pct<threshold and total>0})

    above_count = sum(1 for r in report_data if r['percentage']>=threshold)
    labels      = [r['student'].name for r in report_data]
    percentages = [r['percentage']   for r in report_data]

    return render(request,'attendance/report.html',{
        'report_data':report_data,'query':query,'date_from':date_from,'date_to':date_to,
        'class_id':class_id,'subject_id':subject_id,
        'labels':labels,'percentages':percentages,
        'above_count':above_count,'below_count':len(report_data)-above_count,
        'threshold':threshold,'classes':ClassSection.objects.all(),'subjects':Subject.objects.all(),
    })

def download_excel(request):
    if is_teacher(request):
        records = Attendance.objects.all().select_related('student','subject','timetable_entry'); filename="attendance_report.xlsx"
    elif is_student(request):
        student = get_object_or_404(Student,id=request.session.get('student_id'))
        records = Attendance.objects.filter(student=student).select_related('student','subject','timetable_entry'); filename=f"attendance_{student.roll_number}.xlsx"
    else: return redirect('login')

    wb = openpyxl.Workbook(); sheet = wb.active; sheet.title="Attendance Report"
    headers = ["Student Name","Roll Number","Subject","Lecture","Date","Status","Note"]
    sheet.append(headers)
    for col_num, h in enumerate(headers,1):
        cell = sheet.cell(row=1,column=col_num)
        cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="1F2937"); cell.alignment=Alignment(horizontal="center")

    for r in records:
        lecture_label = "-"
        if r.timetable_entry:
            lecture_label = f"{r.timetable_entry.day} {r.timetable_entry.start_time:%H:%M}-{r.timetable_entry.end_time:%H:%M}"
        sheet.append([r.student.name, r.student.roll_number, r.subject.name if r.subject else "General", lecture_label, str(r.date), r.status, r.note])

    for col in sheet.columns:
        mw = max((len(str(c.value)) for c in col if c.value), default=10)
        sheet.column_dimensions[col[0].column_letter].width = mw+4

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(resp); return resp


def _simple_pdf_response(lines, filename):
    def _escape_pdf_text(value):
        return value.replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')

    # Keep fallback concise so it fits one page in basic PDF mode.
    safe_lines = lines[:45]
    content_lines = ["BT", "/F1 10 Tf", "40 800 Td", "13 TL"]
    for line in safe_lines:
        content_lines.append(f"({_escape_pdf_text(line)}) Tj")
        content_lines.append("T*")
    content_lines.append("ET")
    content_stream = "\n".join(content_lines)

    objects = []

    def add_obj(data):
        objects.append(data.encode('latin-1', errors='replace'))

    add_obj("<< /Type /Catalog /Pages 2 0 R >>")
    add_obj("<< /Type /Pages /Kids [3 0 R] /Count 1 >>")
    add_obj("<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>")
    add_obj(f"<< /Length {len(content_stream.encode('latin-1', errors='replace'))} >>\nstream\n{content_stream}\nendstream")
    add_obj("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")

    pdf = b"%PDF-1.4\n%----\n"
    offsets = [0]
    for idx, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf += f"{idx} 0 obj\n".encode('ascii') + obj + b"\nendobj\n"

    xref_start = len(pdf)
    pdf += f"xref\n0 {len(objects) + 1}\n".encode('ascii')
    pdf += b"0000000000 65535 f \n"
    for off in offsets[1:]:
        pdf += f"{off:010d} 00000 n \n".encode('ascii')

    pdf += f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_start}\n%%EOF".encode('ascii')

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


def download_pdf(request):

    if is_teacher(request):
        records = Attendance.objects.all().select_related('student', 'subject', 'timetable_entry').order_by('-date')
        filename = "attendance_report.pdf"
        title = "Attendance Report"
    elif is_student(request):
        student = get_object_or_404(Student, id=request.session.get('student_id'))
        records = Attendance.objects.filter(student=student).select_related('student', 'subject', 'timetable_entry').order_by('-date')
        filename = f"attendance_{student.roll_number}.pdf"
        title = f"Attendance Report - {student.name}"
    else:
        return redirect('login')

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename={filename}'

        p = canvas.Canvas(response, pagesize=A4)
        page_width, page_height = A4

        y = page_height - 50
        p.setFont("Helvetica-Bold", 14)
        p.drawString(40, y, title)
        y -= 24

        p.setFont("Helvetica-Bold", 10)
        p.drawString(40, y, "Date")
        p.drawString(110, y, "Student")
        p.drawString(240, y, "Subject")
        p.drawString(340, y, "Lecture")
        p.drawString(470, y, "Status")
        y -= 14
        p.line(40, y, page_width - 40, y)
        y -= 14

        p.setFont("Helvetica", 9)
        for r in records:
            lecture_label = "-"
            if r.timetable_entry:
                lecture_label = f"{r.timetable_entry.day[:3]} {r.timetable_entry.start_time:%H:%M}"

            student_name = r.student.name[:22]
            subject_name = (r.subject.name if r.subject else "General")[:16]
            lecture_text = lecture_label[:22]

            p.drawString(40, y, str(r.date))
            p.drawString(110, y, student_name)
            p.drawString(240, y, subject_name)
            p.drawString(340, y, lecture_text)
            p.drawString(470, y, r.status)
            y -= 14

            if y < 60:
                p.showPage()
                y = page_height - 50
                p.setFont("Helvetica-Bold", 10)
                p.drawString(40, y, "Date")
                p.drawString(110, y, "Student")
                p.drawString(240, y, "Subject")
                p.drawString(340, y, "Lecture")
                p.drawString(470, y, "Status")
                y -= 14
                p.line(40, y, page_width - 40, y)
                y -= 14
                p.setFont("Helvetica", 9)

        p.save()
        return response

    except ImportError:
        fallback_lines = [
            title,
            "",
            "Date | Student | Subject | Lecture | Status",
            "-" * 72,
        ]
        for r in records:
            lecture_label = "-"
            if r.timetable_entry:
                lecture_label = f"{r.timetable_entry.day[:3]} {r.timetable_entry.start_time:%H:%M}"
            fallback_lines.append(
                f"{r.date} | {r.student.name[:16]} | {(r.subject.name if r.subject else 'General')[:12]} | {lecture_label[:12]} | {r.status}"
            )

        return _simple_pdf_response(fallback_lines, filename)

@teacher_required
def leaves(request):
    if not is_teacher(request):
        return redirect('login')
    
    status_filter = request.GET.get('status','')
    type_filter = request.GET.get('type','student')  # default to student, student, teacher
    
    qs = LeaveRequest.objects.all().order_by('-applied_on')
     
    if type_filter == 'student':
        qs = qs.filter(student__isnull=False).select_related('student')
    elif type_filter == 'teacher':
        qs = qs.filter(teacher__isnull=False).select_related('teacher')
    else:  # all (fallback)
        qs = qs.select_related('student', 'teacher')
    
    if status_filter:
        qs = qs.filter(status=status_filter)
    
    return render(request, "attendance/leaves.html", {
        "leaves": qs,
        "status_filter": status_filter,
        "type_filter": type_filter,
    })

def apply_leave(request):
    if not is_student(request): return redirect('login')
    student = get_object_or_404(Student, id=request.session.get('student_id'))
    if request.method == "POST":
        fd = request.POST.get('from_date','')
        td = request.POST.get('to_date','')
        reason = request.POST.get('reason','').strip()
        if not fd or not td or not reason:
            messages.error(request,"All fields are required.")
        elif fd > td:
            messages.error(request,"From date cannot be after to date.")
        else:
            LeaveRequest.objects.create(student=student,from_date=fd,to_date=td,reason=reason)
            log_activity(request,'leave',f"Student '{student.name}' applied for leave {fd} to {td}")
            push_notification("New Leave Request",f"{student.name} applied for leave {fd} to {td}",'info')
            messages.success(request,"Leave application submitted!")
            return redirect('student_dashboard')
    return render(request,"attendance/apply_leave.html",{"student":student,"today":str(date.today())})


@teacher_required
def teacher_apply_leave(request):
    teacher = get_object_or_404(Teacher, id=request.session.get('teacher_db_id'))
    if request.method == "POST":
        fd = request.POST.get('from_date','')
        td = request.POST.get('to_date','')
        reason = request.POST.get('reason','').strip()
        if not fd or not td or not reason:
            messages.error(request,"All fields are required.")
        elif fd > td:
            messages.error(request,"From date cannot be after to date.")
        else:
            LeaveRequest.objects.create(teacher=teacher, from_date=fd, to_date=td, reason=reason)
            log_activity(request,'leave',f"Teacher '{teacher.name}' applied for leave {fd} to {td}")
            push_notification("Teacher Leave Request",f"{teacher.name} applied for leave {fd} to {td}",'info')
            messages.success(request,"Leave application submitted!")
            return redirect('home')
    return render(request,"attendance/teacher_apply_leave.html",{"teacher":teacher,"today":str(date.today())})


def my_leaves(request):
    if not is_student(request):
        return redirect('login')

    student = get_object_or_404(Student, id=request.session.get('student_id'))
    leave_requests = LeaveRequest.objects.filter(student=student).order_by('-applied_on')
    return render(request, "attendance/my_leaves.html", {
        "student": student,
        "leave_requests": leave_requests,
    })

@teacher_required
def teacher_my_leaves(request):
    teacher = get_object_or_404(Teacher, id=request.session.get('teacher_db_id'))
    leave_requests = LeaveRequest.objects.filter(teacher=teacher).order_by('-applied_on')
    return render(request, "attendance/teacher_my_leaves.html", {
        "teacher": teacher,
        "leave_requests": leave_requests,
    })

@teacher_required
def review_leave(request, id, action):
    leave = get_object_or_404(LeaveRequest, id=id)
    if action not in ['approve','reject']:
        return redirect('leaves')
    note = request.POST.get('note','')
    from django.utils import timezone as tz
    leave.status       = 'Approved' if action=='approve' else 'Rejected'
    leave.reviewed_on  = tz.now()
    leave.reviewer_note = note
    leave.save()
    msg_type = 'success' if action=='approve' else 'warning'
    push_notification(
        f"Leave {leave.status}",
        f"Your leave from {leave.from_date} to {leave.to_date} was {leave.status.lower()}.",
        msg_type, student=leave.student
    )
    log_activity(request,'leave',f"Leave for '{leave.student.name}' {leave.status}")
    messages.success(request,f"Leave {leave.status}!")
    return redirect('leaves')


def notifications(request):
    if is_teacher(request):
        tid  = request.session.get('teacher_db_id')
        notifs = Notification.objects.filter(recipient_teacher_id=tid).order_by('-created_at')[:50] if tid else []
    elif is_student(request):
        sid    = request.session.get('student_id')
        notifs = Notification.objects.filter(recipient_student_id=sid).order_by('-created_at')[:50]
    else: return redirect('login')
    Notification.objects.filter(id__in=[n.id for n in notifs]).update(is_read=True)
    return render(request,"attendance/notifications.html",{"notifications":notifs})

def mark_notification_read(request, id):
    Notification.objects.filter(id=id).update(is_read=True)
    return redirect('notifications')


@admin_required
def activity_logs(request):
    logs = ActivityLog.objects.all()[:200]
    return render(request,"attendance/activity_logs.html",{"logs":logs})


def student_dashboard(request):
    if not is_student(request): return redirect('login')
    student = get_object_or_404(Student, id=request.session.get('student_id'))
    selected_date  = request.GET.get('date')
    selected_month = request.GET.get('month')

    att_q = Attendance.objects.filter(student=student)
    if selected_date:   att_q = att_q.filter(date=selected_date)
    elif selected_month:
        yr,mn = selected_month.split("-")
        att_q = att_q.filter(date__year=yr,date__month=mn)

    all_attendance = att_q.order_by('-date')
    present,absent,total,pct = student_attendance_stats(student)
    can_miss, _ = attendance_prediction(student)

    today = date.today()
    cal_data = build_student_calendar_data(student, request.GET.get('cal_month'), today=today)

    pending_leaves = LeaveRequest.objects.filter(student=student,status='Pending').count()
    unread = get_unread_count(request)

    threshold  = getattr(django_settings,'ATTENDANCE_THRESHOLD',75)
    return render(request,'attendance/student_dashboard.html',{
        'student':student,'all_attendance':all_attendance,'present':present,'absent':absent,
        'total':total,'percentage':pct,'can_miss':can_miss,'selected_date':selected_date or today,
        'today':today,'greeting':get_greeting(),'threshold':threshold,
        'calendar_days':cal_data['calendar_days'],'calendar_blank':cal_data['calendar_blank'],'cal_month':cal_data['cal_month'],
        'cal_month_value': cal_data['cal_month_value'],
        'prev_cal_month': cal_data['prev_cal_month'],
        'next_cal_month': cal_data['next_cal_month'],
        'pending_leaves':pending_leaves,'unread':unread,
    })


def student_calendar_data(request):
    if not is_student(request):
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    student = get_object_or_404(Student, id=request.session.get('student_id'))
    cal_data = build_student_calendar_data(student, request.GET.get('cal_month'), today=date.today())

    return JsonResponse({
        'cal_month': cal_data['cal_month'],
        'cal_month_value': cal_data['cal_month_value'],
        'prev_cal_month': cal_data['prev_cal_month'],
        'next_cal_month': cal_data['next_cal_month'],
        'calendar_blank': cal_data['calendar_blank'],
        'calendar_days': [
            {
                'date': d['date'].strftime('%Y-%m-%d'),
                'day': d['date'].day,
                'status': d['status'],
                'holiday': d['holiday'],
                'is_today': d['is_today'],
            }
            for d in cal_data['calendar_days']
        ],
    })


def student_profile(request):
    if not is_student(request):
        return redirect('login')

    student = get_object_or_404(Student, id=request.session.get('student_id'))
    return render(request, 'attendance/student_profile.html', {
        'student': student,
        'greeting': get_greeting(),
        'unread': get_unread_count(request),
    })


def teacher_profile(request):
    if not is_teacher(request):
        return redirect('login')

    teacher = None
    teacher_id = request.session.get('teacher_db_id')
    if teacher_id:
        teacher = Teacher.objects.filter(id=teacher_id).first()

    return render(request, 'attendance/teacher_profile.html', {
        'teacher': teacher,
        'teacher_name': request.session.get('teacher_name', 'Teacher'),
        'greeting': get_greeting(),
        'unread': get_unread_count(request),
    })


def student_timetable(request):
    if not is_student(request):
        return redirect('login')

    student = get_object_or_404(Student, id=request.session.get('student_id'))
    selected_class = student.class_section
    auto_selected = False
    DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    schedule = {day: [] for day in DAYS}

    if not selected_class:
        fallback_classes = ClassSection.objects.filter(timetable__isnull=False).distinct()
        if fallback_classes.count() == 1:
            selected_class = fallback_classes.first()
            auto_selected = True

    if selected_class:
        entries = Timetable.objects.filter(class_section=selected_class).select_related('subject', 'teacher')
        for day in DAYS:
            schedule[day] = entries.filter(day=day).order_by('start_time')

    return render(request, 'attendance/student_timetable.html', {
        'student': student,
        'selected_class': selected_class,
        'auto_selected': auto_selected,
        'schedule': schedule,
        'days': DAYS,
    })


@teacher_required
def admin_panel(request):
    today     = date.today()
    query     = request.GET.get('q','')
    threshold = getattr(django_settings,'ATTENDANCE_THRESHOLD',75)

    total_students = Student.objects.count()
    total_teachers = Teacher.objects.count()
    present_today  = Attendance.objects.filter(date=today,status='Present',subject=None).count()
    absent_today   = Attendance.objects.filter(date=today,status='Absent', subject=None).count()
    total_records  = Attendance.objects.count()
    pending_leaves = LeaveRequest.objects.filter(status='Pending').count()
    pending_leave_requests = LeaveRequest.objects.filter(status='Pending').select_related('student')[:10]

    qs = Student.objects.all()
    if query: qs = qs.filter(Q(name__icontains=query)|Q(roll_number__icontains=query))

    student_rows=[]; low_students=[]; chart_data=[]
    for student in Student.objects.all():
        p,a,t,pct = student_attendance_stats(student)
        chart_data.append({'name':student.name,'percentage':pct})
        if pct<threshold and t>0: low_students.append({'student':student,'percentage':pct,'present':p,'total':t})

    for student in qs:
        att = Attendance.objects.filter(student=student,date=today,subject=None).first()
        _,_,_,pct = student_attendance_stats(student)
        student_rows.append({'student':student,'today_status':att.status if att else None,'percentage':pct})

    recent_logs = ActivityLog.objects.all()[:8]
    db_engine   = django_settings.DATABASES['default']['ENGINE'].split('.')[-1].upper()

    return render(request,'attendance/admin_panel.html',{
        'total_students':total_students,'total_teachers':total_teachers,
        'present_today':present_today,'absent_today':absent_today,
        'total_records':total_records,'pending_leaves':pending_leaves,
        'pending_leave_requests':pending_leave_requests,
        'student_rows':student_rows,'chart_data':chart_data[:10],
        'low_students':low_students,'low_attendance_count':len(low_students),
        'recent_logs':recent_logs,'db_engine':db_engine,
        'debug_mode':getattr(django_settings,'DEBUG',False),
        'query':query,'threshold':threshold,
    })

@teacher_required
def timetable(request):
    classes   = ClassSection.objects.all()
    class_id  = request.GET.get('class_id', '')
    selected_class = None
    schedule = {}

    DAYS = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday']

    if class_id:
        try:
            selected_class = ClassSection.objects.get(id=class_id)
            entries = Timetable.objects.filter(class_section=selected_class).select_related('subject','teacher')
            for day in DAYS:
                schedule[day] = entries.filter(day=day).order_by('start_time')
        except ClassSection.DoesNotExist:
            pass
    elif classes.exists():
        selected_class = classes.first()
        entries = Timetable.objects.filter(class_section=selected_class).select_related('subject','teacher')
        for day in DAYS:
            schedule[day] = entries.filter(day=day).order_by('start_time')
        class_id = str(selected_class.id)

    subjects = Subject.objects.filter(class_section=selected_class) if selected_class else Subject.objects.all()
    teachers = Teacher.objects.all()

    return render(request, 'attendance/timetable.html', {
        'classes':        classes,
        'selected_class': selected_class,
        'class_id':       class_id,
        'schedule':       schedule,
        'days':           DAYS,
        'subjects':       subjects,
        'teachers':       teachers,
    })


@teacher_required
def add_timetable(request):
    if request.method == 'POST':
        class_id   = request.POST.get('class_section')
        subject_id = request.POST.get('subject')
        teacher_id = request.POST.get('teacher')
        day        = request.POST.get('day')
        start      = request.POST.get('start_time')
        end        = request.POST.get('end_time')

        if not all([class_id, subject_id, day, start, end]):
            messages.error(request, 'All fields except teacher are required.')
            return redirect(f'/timetable/?class_id={class_id}')

        # Check for time conflict
        conflict = Timetable.objects.filter(
            class_section_id=class_id, day=day
        ).filter(
            start_time__lt=end, end_time__gt=start
        ).first()

        if conflict:
            messages.error(request, f"Time conflict with {conflict.subject.name} ({conflict.start_time:%H:%M}–{conflict.end_time:%H:%M})")
            return redirect(f'/timetable/?class_id={class_id}')

        Timetable.objects.create(
            class_section_id=class_id,
            subject_id=subject_id,
            teacher_id=teacher_id or None,
            day=day,
            start_time=start,
            end_time=end,
        )
        messages.success(request, 'Lecture added to timetable!')
        return redirect(f'/timetable/?class_id={class_id}')
    return redirect('timetable')


@teacher_required
def delete_timetable(request, id):
    entry = get_object_or_404(Timetable, id=id)
    class_id = entry.class_section_id
    entry.delete()
    messages.success(request, 'Lecture removed from timetable.')
    return redirect(f'/timetable/?class_id={class_id}')

@teacher_required
def send_low_attendance_emails(request):
    threshold = getattr(django_settings,'ATTENDANCE_THRESHOLD',75)
    low_list  = []
    for student in Student.objects.all():
        p,a,t,pct = student_attendance_stats(student)
        if t>0 and pct<threshold: low_list.append((student,p,t,pct))

    if not low_list:
        messages.info(request,f"No students below {threshold}%."); return redirect('report')

    sent=0; missing=0; failed=0
    for student,present,total,pct in low_list:
        if not student.email: missing+=1; continue
        try:
            send_mail("⚠ Low Attendance Alert",
                f"Dear {student.name},\n\nYour attendance: {pct}%\nPresent: {present} / {total}\nRequired: {threshold}%\n\nPlease attend regularly.\n\nEduTrack",
                django_settings.DEFAULT_FROM_EMAIL,[student.email],fail_silently=False)
            sent+=1
        except: failed+=1

    log_activity(request,'email',f"Sent low attendance alerts: {sent} sent, {failed} failed")
    if sent: messages.success(request,f"Alerts sent to {sent} student(s).")
    elif failed: messages.error(request,"Email failed. Check SMTP in .env")
    else: messages.warning(request,f"{len(low_list)} students below threshold but no emails registered.")
    return redirect('report')
